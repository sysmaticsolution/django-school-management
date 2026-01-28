"""
Excel Report Generation for School Management System.
Uses openpyxl for Excel creation.
"""
import io
from datetime import date
from decimal import Decimal

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelReportMixin:
    """Common functionality for Excel reports."""
    
    @staticmethod
    def apply_header_style(cell):
        """Apply header styling to a cell."""
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    @staticmethod
    def apply_cell_style(cell, align='left'):
        """Apply standard cell styling."""
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
    
    @staticmethod
    def auto_adjust_columns(worksheet):
        """Auto-adjust column widths based on content."""
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width


class StudentListExcel(ExcelReportMixin):
    """Generate student list Excel."""
    
    def __init__(self, section=None, filters=None):
        self.section = section
        self.filters = filters or {}
    
    def generate(self):
        """Generate student list Excel."""
        from apps.students.models import Student
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Student List"
        
        # Query students
        queryset = Student.objects.select_related('current_section', 'current_section__standard')
        if self.section:
            queryset = queryset.filter(current_section=self.section)
        queryset = queryset.filter(is_active=True).order_by('current_section', 'roll_number')
        
        # Headers
        headers = [
            'S.No.', 'Admission No.', 'Student Name', 'Class', 'Section',
            'Roll No.', 'Gender', 'DOB', 'Father Name', 'Phone', 'Category'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.apply_header_style(cell)
        
        # Data rows
        for row_num, student in enumerate(queryset, 2):
            data = [
                row_num - 1,
                student.admission_number,
                student.full_name,
                student.current_section.standard.name if student.current_section else '-',
                student.current_section.name if student.current_section else '-',
                student.roll_number or '-',
                student.get_gender_display() if hasattr(student, 'get_gender_display') else student.gender,
                student.date_of_birth.strftime('%d/%m/%Y') if student.date_of_birth else '-',
                student.father_name or '-',
                student.phone or '-',
                student.category or '-'
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                self.apply_cell_style(cell, 'center' if col in [1, 6, 7] else 'left')
        
        self.auto_adjust_columns(ws)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class FeeDefaultersExcel(ExcelReportMixin):
    """Generate fee defaulters Excel report."""
    
    def __init__(self, academic_year=None):
        self.academic_year = academic_year
    
    def generate(self):
        """Generate fee defaulters Excel."""
        from apps.fees.models import StudentFee
        from apps.students.models import Student
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Defaulters"
        
        # Query pending fees
        queryset = StudentFee.objects.filter(
            status__in=['pending', 'partial']
        ).select_related(
            'student', 'student__current_section', 'student__current_section__standard',
            'fee_structure', 'fee_structure__category'
        ).order_by('student__current_section', 'student__roll_number')
        
        if self.academic_year:
            queryset = queryset.filter(fee_structure__academic_year=self.academic_year)
        
        # Headers
        headers = [
            'S.No.', 'Admission No.', 'Student Name', 'Class', 'Phone',
            'Fee Type', 'Total Amount', 'Paid', 'Balance', 'Due Date', 'Status'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.apply_header_style(cell)
        
        # Data rows
        for row_num, fee in enumerate(queryset, 2):
            data = [
                row_num - 1,
                fee.student.admission_number,
                fee.student.full_name,
                f"{fee.student.current_section.standard.name} - {fee.student.current_section.name}",
                fee.student.phone or fee.student.father_phone or '-',
                fee.fee_structure.category.name,
                float(fee.final_amount),
                float(fee.paid_amount),
                float(fee.balance),
                fee.due_date.strftime('%d/%m/%Y') if fee.due_date else '-',
                fee.get_status_display()
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                if col in [7, 8, 9]:
                    cell.number_format = '₹#,##0.00'
                self.apply_cell_style(cell, 'center' if col in [1, 4, 11] else 'left')
        
        # Summary row
        total_row = len(queryset) + 2
        ws.cell(row=total_row, column=6, value="TOTAL:").font = Font(bold=True)
        
        total_amount = sum(float(f.final_amount) for f in queryset)
        total_paid = sum(float(f.paid_amount) for f in queryset)
        total_balance = sum(float(f.balance) for f in queryset)
        
        for col, val in [(7, total_amount), (8, total_paid), (9, total_balance)]:
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.font = Font(bold=True)
            cell.number_format = '₹#,##0.00'
        
        self.auto_adjust_columns(ws)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class AttendanceSummaryExcel(ExcelReportMixin):
    """Generate attendance summary Excel."""
    
    def __init__(self, section, month, year):
        self.section = section
        self.month = month
        self.year = year
    
    def generate(self):
        """Generate attendance summary Excel."""
        from apps.attendance.models import AttendanceSummary
        from apps.students.models import Student
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {self.month}-{self.year}"
        
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1, 
            value=f"Attendance Summary - {self.section.standard.name} ({self.section.name}) - {self.month}/{self.year}")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['S.No.', 'Roll No.', 'Student Name', 'Working Days', 'Present', 'Absent', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self.apply_header_style(cell)
        
        # Get students
        students = Student.objects.filter(
            current_section=self.section,
            is_active=True
        ).order_by('roll_number')
        
        row_num = 4
        for idx, student in enumerate(students, 1):
            summary = AttendanceSummary.objects.filter(
                student=student,
                month=self.month,
                year=self.year
            ).first()
            
            if summary:
                working_days = summary.total_working_days
                present = summary.present_days
                absent = summary.absent_days
                percentage = (present / working_days * 100) if working_days > 0 else 0
            else:
                working_days = present = absent = 0
                percentage = 0
            
            data = [idx, student.roll_number or '-', student.full_name, 
                    working_days, present, absent, f"{percentage:.1f}%"]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                self.apply_cell_style(cell, 'center' if col != 3 else 'left')
                
                # Color code percentage
                if col == 7 and percentage < 75:
                    cell.font = Font(color='FF0000', bold=True)
            
            row_num += 1
        
        self.auto_adjust_columns(ws)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class ExamResultsExcel(ExcelReportMixin):
    """Generate exam results Excel."""
    
    def __init__(self, exam, section=None):
        self.exam = exam
        self.section = section
    
    def generate(self):
        """Generate exam results Excel."""
        from apps.examinations.models import ExamResult, ReportCard
        from apps.academics.models import Subject
        from apps.students.models import Student
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Results - {self.exam.name[:20]}"
        
        # Get subjects for this exam
        subjects = Subject.objects.filter(
            exam_results__exam=self.exam
        ).distinct().order_by('name')
        
        # Get students
        students_query = Student.objects.filter(is_active=True)
        if self.section:
            students_query = students_query.filter(current_section=self.section)
        students = students_query.order_by('current_section', 'roll_number')
        
        # Create headers
        headers = ['S.No.', 'Roll No.', 'Student Name', 'Class']
        for subject in subjects:
            headers.append(subject.name[:15])
        headers.extend(['Total', 'Percentage', 'Grade', 'Rank', 'Result'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.apply_header_style(cell)
        
        # Data rows
        row_num = 2
        for idx, student in enumerate(students, 1):
            try:
                report_card = ReportCard.objects.get(student=student, exam=self.exam)
            except ReportCard.DoesNotExist:
                continue
            
            row_data = [
                idx, 
                student.roll_number or '-', 
                student.full_name,
                f"{student.current_section.standard.name} - {student.current_section.name}" if student.current_section else '-'
            ]
            
            # Subject marks
            for subject in subjects:
                try:
                    result = ExamResult.objects.get(student=student, exam=self.exam, subject=subject)
                    row_data.append(result.marks_obtained)
                except ExamResult.DoesNotExist:
                    row_data.append('-')
            
            # Summary
            row_data.extend([
                f"{report_card.marks_obtained}/{report_card.total_marks}",
                f"{report_card.percentage:.1f}%",
                report_card.overall_grade,
                report_card.rank or '-',
                'PASS' if report_card.percentage >= 33 else 'FAIL'
            ])
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                self.apply_cell_style(cell, 'center' if col != 3 else 'left')
            
            row_num += 1
        
        self.auto_adjust_columns(ws)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# =============================================================================
# EXCEL VIEW HELPERS
# =============================================================================

def download_student_list_excel(request, section_id=None):
    """Download student list Excel."""
    from apps.academics.models import Section
    
    section = Section.objects.get(pk=section_id) if section_id else None
    excel = StudentListExcel(section=section)
    buffer = excel.generate()
    
    filename = f"student_list_{section.standard.name}_{section.name}.xlsx" if section else "student_list_all.xlsx"
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def download_fee_defaulters_excel(request):
    """Download fee defaulters Excel."""
    excel = FeeDefaultersExcel()
    buffer = excel.generate()
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="fee_defaulters.xlsx"'
    return response


def download_attendance_excel(request, section_id, month, year):
    """Download attendance summary Excel."""
    from apps.academics.models import Section
    
    section = Section.objects.get(pk=section_id)
    excel = AttendanceSummaryExcel(section, int(month), int(year))
    buffer = excel.generate()
    
    filename = f"attendance_{section.standard.name}_{section.name}_{month}_{year}.xlsx"
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def download_exam_results_excel(request, exam_id, section_id=None):
    """Download exam results Excel."""
    from apps.examinations.models import Exam
    from apps.academics.models import Section
    
    exam = Exam.objects.get(pk=exam_id)
    section = Section.objects.get(pk=section_id) if section_id else None
    
    excel = ExamResultsExcel(exam, section=section)
    buffer = excel.generate()
    
    filename = f"results_{exam.name}_{section.name if section else 'all'}.xlsx"
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
